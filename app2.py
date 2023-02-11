import openpyxl
import re
from openpyxl.worksheet.worksheet import Worksheet

# workbook = openpyxl.Workbook()
workbook = openpyxl.load_workbook(filename="data.xlsx")
sheet = workbook.active
last_column = "AD"
print(workbook)

def get_words(row):
    _words = []
    for item in row[9:]:
        _value = item.value
        if not _value:
            continue
        _word = re.findall(r"\w+", str(_value))
        if len(_word) > 0:
            _words.append(_value)
    if _words:
        print(_words)
    return _words

def main(row_no, sheet: Worksheet):
    for idx, row in enumerate(sheet.rows):
        currnet_row = idx + 1
        if idx == 0 or row_no < currnet_row:
            continue
        max_row = sheet.max_row
        words = get_words(row)
        if len(words) > 0:
            sheet.move_range(f"A{currnet_row + 1}:AD{max_row + 1}", rows=len(words), cols=0)
            for cell in row[9:]:
                cell.value = None
            for jdx, word in enumerate(words):
                sheet[f"I{currnet_row + jdx + 1}"] = word
            break
        if currnet_row == max_row:
            return True
    main(currnet_row, sheet)

main(1, sheet)
workbook.save("result.xlsx")